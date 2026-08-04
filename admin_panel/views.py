import logging
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import render

from .models import Book, OzonTemplate

logger = logging.getLogger(__name__)




BOOK_TYPE_OZON_MAPPING = {
    'printed book': 'Печатная книга',
    'second': 'Second-hand книга',
    'bookinist': 'Букинистика',
    'print_on_demand': 'Печать по требованию',
}


BOOK_TYPE_DISPLAY_MAPPING = {
    'printed book': 'Печатная книга',
    'second': 'Б/У',
    'bookinist': 'Букинистика',
    'print_on_demand': 'Печать по требованию',
}


def _ozon_book_type(book: Book) -> str:
    return BOOK_TYPE_OZON_MAPPING.get(book.book_type, 'Печатная книга')


def _ozon_book_type_display(book: Book) -> str:
    return BOOK_TYPE_DISPLAY_MAPPING.get(book.book_type, 'Печатная книга')


AVITO_BOOK_TYPE_MAPPING = {
    'printed book': 'Книги',
    'second': 'Second-hand книги',
    'bookinist': 'Букинистика',
    'print_on_demand': 'Печать по требованию',
}


def _avito_book_type(book: Book) -> str:
    return AVITO_BOOK_TYPE_MAPPING.get(book.book_type, 'Книги')


def _avito_condition(book: Book) -> str:
    return 'Б/у'


AVITO_GENRE_MAPPING = {
    'detective': 'Детективы',
    'fantastic': 'Фантастика',
    'fantasy': 'Фэнтези',
    'antic': 'Антиутопии',
    'adventure': 'Приключения',
    'classic': 'Классическая проза',
    'modern': 'Современная проза',
    'history': 'Историческая проза',
    'romance': 'Любовная проза',
    'drama': 'Пьесы и драматургия',
    'poetry': 'Поэзия',
    'epic_folklore': 'Фольклор и мифология',
    'horror': 'Ужасы и триллеры',
    'thriller': 'Ужасы и триллеры',
    'boevic': 'Боевики',
    'satire': 'Юмор',
}


def _avito_genre(book: Book) -> str:
    return book.genre or ''


def _avito_photos_links(book: Book) -> str:
    if not book.photos:
        return ''
    return ' | '.join(f'{settings.MEDIA_BASE_URL}{photo}' for photo in book.photos)


def _avito_photos_names(book: Book) -> str:
    if not book.photos:
        return ''
    return ' | '.join(photo.split('/')[-1] for photo in book.photos)


AVITO_FIELD_MAPPING = {
    'уникальный идентификатор объявления': lambda book: book.sku or f'book_{book.id}' if book.id else '',
    'начало размещения': lambda book: '',
    'окончание размещения': lambda book: '',
    'способ размещения': lambda book: '',
    'услуга продвижения': lambda book: '',
    'номер объявления на авито': lambda book: '',
    'контактное лицо': lambda book: '',
    'номер телефона': lambda book: '',
    'адрес': lambda book: '',
    'широта': lambda book: '',
    'долгота': lambda book: '',
    'идентификатор адреса': lambda book: '',
    'название объявления': lambda book: (book.title or '')[:50],
    'описание объявления': lambda book: (book.description or '')[:7500],
    'ссылки на фото': _avito_photos_links,
    'названия фото': _avito_photos_names,
    'ссылка на видео': lambda book: '',
    'способ связи': lambda book: 'По телефону и в сообщениях',
    'настройка цены целевого действия': lambda book: '',
    'настройка цены целевого действия: автоматическая': lambda book: '',
    'настройка цены целевого действия: ручная': lambda book: '',
    'категория': lambda book: 'Книги и журналы',
    'интернет звонки': lambda book: '',
    'устройства для приёма звонков': lambda book: '',
    'способ доставки': lambda book: '',
    'вес (для доставки)': lambda book: str(round(float(book.weight) / 1000, 3)) if book.weight else '',
    'длина (для доставки)': lambda book: str(round(float(book.length) / 10, 1)) if book.length else '',
    'высота (для доставки)': lambda book: str(round(float(book.height) / 10, 1)) if book.height else '',
    'ширина (для доставки)': lambda book: str(round(float(book.width) / 10, 1)) if book.width else '',
    'возвраты': lambda book: '',
    'цена': lambda book: str(int(float(book.price))) if book.price else '',
    'вид товара': lambda book: 'Книги',
    'вид объявления': lambda book: 'Товар приобретен на продажу',
    'состояние': _avito_condition,
    'url видеофайла': lambda book: '',
    'вид книги': lambda book: 'Художественная литература',
    'жанр': _avito_genre,
    'автор': lambda book: book.author or '',
    'популярная серия': lambda book: book.series or '',
}


def _tnved_code(book: Book) -> str:
    return book.tnved_code or '4901100000 - Книги, брошюры, листовки и аналогичные печатные издания в виде отдельных листов, сфальцованные или несфальцованные'


def _format_dimensions_cm(book: Book) -> str:
    if book.length and book.width and book.height:
        try:
            l_cm = round(float(book.length) / 10, 1)
            w_cm = round(float(book.width) / 10, 1)
            h_cm = round(float(book.height) / 10, 1)
            return f"{l_cm} x {w_cm} x {h_cm}"
        except (ValueError, TypeError):
            pass
    return ''


def _format_dimensions_mm(book: Book) -> str:
    if book.length and book.width and book.height:
        try:
            return f"{int(float(book.length))} x {int(float(book.width))} x {int(float(book.height))}"
        except (ValueError, TypeError):
            pass
    return ''


def dashboard(request):
    return render(request, 'admin_panel/dashboard.html')
def products(request):
    return render(request, 'admin_panel/products.html')
def orders(request):
    return render(request, 'admin_panel/orders.html')
def customers(request):
    return render(request, 'admin_panel/customers.html')
def export_books_to_ozon_template(request):
    """
    Экспорт выбранных книг в шаблон Ozon.
    Загружает активный шаблон Excel из модели OzonTemplate и заполняет
    данными книг с автоматическим маппингом всех 78 колонок шаблона.
    Особенности:
        - Маппинг всех колонок шаблона на поля модели Book
        - Автоматическое определение направления (колонки 47-72)
        - Форматирование размеров в см и мм
        - Тип* и Тип книги определяются из book_type
    """
    if hasattr(request, 'ozon_export_queryset'):
        books = request.ozon_export_queryset
    else:
        books = Book.objects.select_related('category').all()
    template = OzonTemplate.objects.filter(is_active=True).first()
    if not template:
        return HttpResponse(
            "Ошибка: Не найден активный шаблон Ozon. Загрузите шаблон через админку.",
            content_type='text/plain; charset=utf-8',
            status=400
        )
    template_path = os.path.join(settings.MEDIA_ROOT, template.file.name)
    if not os.path.exists(template_path):
        return HttpResponse(
            f"Ошибка: Файл шаблона не найден: {template.file.name}",
            content_type='text/plain; charset=utf-8',
            status=404
        )
    media_base_url = getattr(settings, 'MEDIA_BASE_URL',
                              request.build_absolute_uri(settings.MEDIA_URL))
    try:
        wb = load_workbook(template_path)
        if 'Шаблон' not in wb.sheetnames:
            return HttpResponse(
                "Ошибка: В шаблоне не найден лист 'Шаблон'",
                content_type='text/plain; charset=utf-8',
                status=400
            )
        ws = wb['Шаблон']
        header_row = 2
        headers = {}
        for col_num in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col_num).value
            if cell_value:
                header_clean = ' '.join(str(cell_value).replace('\n', ' ').split()).lower()
                headers[header_clean] = col_num
        logger.info(f'Ozon template headers: {list(headers.keys())}')
        isbn_cols = {k: v for k, v in headers.items() if 'isbn' in k.lower()}
        logger.info(f'ISBN columns found: {isbn_cols}')
        field_mapping = {
            'артикул*': lambda book: book.sku or '',
            'название товара': lambda book: book.title or '',
            'цена, руб.*': lambda book: float(book.price) if book.price else '',
            'цена до скидки, руб.': lambda book: float(book.old_price) if book.old_price else '',
            'ндс, %*': lambda book: int(book.vat_rate) if book.vat_rate else 0,
            #'sku': lambda book: book.sku or '',
            'штрихкод (серийный номер / ean)': lambda book: '',  # book.isbn or '',
            'isbn': lambda book: book.isbn or '',
            'isbn*': lambda book: book.isbn or '',
            'вес в упаковке, г*': lambda book: int(float(book.weight)) if book.weight else '',
            'ширина упаковки, мм*': lambda book: int(float(book.width)) if book.width else '',  # убрал  / 10 
            'высота упаковки, мм*': lambda book: int(float(book.height)) if book.height else '',  # убрал  / 10
            'длина упаковки, мм*': lambda book: int(float(book.length)) if book.length else '',  # убрал  / 10

            'ссылка на главное фото*': lambda book: (
                f'{media_base_url}{book.photos[0]}' if book.photos else ''
            ),
            'ссылки на дополнительные фото': lambda book: (
                ', '.join(f'{media_base_url}{photo}' for photo in book.photos[1:])
                if len(book.photos) > 1 else ''
            ),
            'артикул фото': lambda book: book.sku or '',
            'автор на обложке*': lambda book: (
                book.author_oblozh if book.author_oblozh else (book.author or '')
            ),
            'автор': lambda book: book.author or '',
            'тип обложки': lambda book: book.get_cover_type_display() or '',            
            #'тип книги': lambda book: _ozon_book_type_display(book),
            'тип книги': lambda book: _ozon_book_type(book),
            'тип*': lambda book: 'Печатная книга',
            'бренд*': lambda book: book.publisher or 'Нет бренда',
            'тн вэд коды еаэс': _tnved_code,
            'тн вэд коды еаэс*': _tnved_code,
            'тнвэд': _tnved_code,
            'тнвэд*': _tnved_code,
            'код тн вэд': _tnved_code,
            'код тн вэд*': _tnved_code,
            'направление*': lambda book: book.get_genre_display() or '',
            'целевая аудитория литературы': lambda book: book.get_target_audience_display() or '',
            '#хештеги': lambda book: book.hashtags or '',
            'аннотация': lambda book: book.description or '',
            'иллюстратор': lambda book: book.illustrator or '',
            'переводчик': lambda book: book.translator or '',
            'издательство': lambda book: book.publisher or '',
            'серия': lambda book: book.series or '',
            'год выпуска': lambda book: book.publication_year or '',
            'тип бумаги в книге': lambda book: book.get_paper_type_display() or '',
            'язык издания': lambda book: book.get_language_display() or 'Русский',
            'количество страниц': lambda book: book.pages or '',
            #'размер упаковки (длина х ширина х высота), см': lambda book: _format_dimensions_cm(book),
            #'размеры, мм': lambda book: _format_dimensions_mm(book),
            'вес товара, г': lambda book: int(float(book.weight)) if book.weight else '',
            'сохранность книги': lambda book: book.get_condition_display() or '',
            'возрастные ограничения': lambda book: book.get_age_restrictions_display() or '',
            'признак 18+': lambda book:book.is_adult
        }
        current_row = 5
        for idx, book in enumerate(books, 1):
            ws.cell(row=current_row, column=1).value = idx
            for header_name, col_num in headers.items():
                mapper = field_mapping.get(header_name) or field_mapping.get(header_name.rstrip('*'))
                if mapper:
                    try:
                        value = mapper(book)
                        ws.cell(row=current_row, column=col_num).value = value
                    except Exception as e:
                        logger.warning(
                            f"Ошибка при заполнении '{header_name}' "
                            f"для книги {book.id} ({book.sku}): {e}"
                        )
                        ws.cell(row=current_row, column=col_num).value = ''
            current_row += 1
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="ozon_export_{len(books)}_books.xlsx"'
        wb.save(response)
        return response
    except Exception as e:
        logger.error(f"Ошибка при экспорте в шаблон Ozon: {e}")
        return HttpResponse(
            f"Ошибка при обработке шаблона: {str(e)}",
            content_type='text/plain; charset=utf-8',
            status=500
        )


def export_books_to_avito_template(request):
    """
    Экспорт выбранных книг в шаблон Avito.
    Загружает активный шаблон Avito (по признаку в имени файла/названии) и заполняет
    данными книг с использованием AVITO_FIELD_MAPPING.
    Особенности:
        - Использует лист «Объявления» вместо «Шаблон»
        - Маппинг заголовков Avito через AVITO_FIELD_MAPPING
        - Обработка ошибок аналогична Ozon-экспорту
    """
    if hasattr(request, 'avito_export_queryset'):
        books = request.avito_export_queryset
    else:
        books = Book.objects.select_related('category').all()
    template = OzonTemplate.objects.filter(is_active=True, file__icontains='avito').first()
    if not template:
        return HttpResponse(
            "Ошибка: Не найден активный шаблон Avito. Загрузите шаблон через админку (в имени файла/названии должно быть «avito»).",
            content_type='text/plain; charset=utf-8',
            status=400
        )
    template_path = os.path.join(settings.MEDIA_ROOT, template.file.name)
    if not os.path.exists(template_path):
        return HttpResponse(
            f"Ошибка: Файл шаблона не найден: {template.file.name}",
            content_type='text/plain; charset=utf-8',
            status=404
        )
    media_base_url = getattr(settings, 'MEDIA_BASE_URL', request.build_absolute_uri(settings.MEDIA_URL))
    try:
        wb = load_workbook(template_path)
        if 'Объявления' not in wb.sheetnames:
            return HttpResponse(
                "Ошибка: В шаблоне отсутствует лист «Объявления».",
                content_type='text/plain; charset=utf-8',
                status=400
            )
        ws = wb['Объявления']
        # Определяем заголовки из 2-й строки
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=2, column=col).value
            if cell_value:
                header_clean = ' '.join(str(cell_value).replace('\n', ' ').split()).lower()
                headers.append((col, header_clean))
        # Заполняем данные
        for row_num, book in enumerate(books, 3):
            for col_num, header_clean in headers:
                if header_clean in AVITO_FIELD_MAPPING:
                    value = AVITO_FIELD_MAPPING[header_clean](book)
                else:
                    value = ''
                ws.cell(row=row_num, column=col_num).value = value
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="avito_export.xlsx"'
        wb.save(response)
        return response
    except Exception as e:
        logger.error(f"Ошибка при экспорте в шаблон Avito: {e}")
        return HttpResponse(
            f"Ошибка при обработке шаблона: {str(e)}",
            content_type='text/plain; charset=utf-8',
            status=500
        )


def export_books_to_excel(request):
    """
    Экспорт выбранных книг в стандартный Excel файл.
    Создает Excel файл (.xlsx) со всеми данными книг в табличном формате.
    Используется библиотека openpyxl для работы с Excel.
    Args:
        request: HTTP запрос с атрибутом excel_export_queryset (queryset книг)
    Returns:
        HttpResponse с Excel файлом для скачивания
    Особенности:
        - 25 колонок с полной информацией о книгах
        - Форматированные заголовки (жирный шрифт, выравнивание)
        - Автоматическая ширина колонок (макс. 50 символов)
        - Преобразование Decimal в float для корректного отображения
    """
    # Получить книги из запроса или все книги, если не указано
    if hasattr(request, 'excel_export_queryset'):
        books = request.excel_export_queryset
    else:
        books = Book.objects.select_related('category').all()
    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Книги"
    # Заголовки колонок - все основные поля книги
    headers = [
        'ID', 'Артикул', 'Название', 'Автор', 'Автор на обложке', 'Жанр',
        'Издательство', 'Серия', 'Год издания', 'Язык', 'Сохранность',
        'Тип переплёта', 'Страниц', 'ISBN', 'Цена', 'Старая цена',
        'НДС', 'Остаток', 'Вес (г)', 'Длина (мм)', 'Ширина (мм)',
        'Высота (мм)', 'Категория', 'Источник', 'Дата создания'
    ]
    # Записываем заголовки в первую строку с форматированием
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)  # Жирный шрифт
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Выравнивание
    # Записываем данные книг начиная со второй строки
    for row_num, book in enumerate(books, 2):
        # Основная информация
        ws.cell(row=row_num, column=1).value = book.id
        ws.cell(row=row_num, column=2).value = book.sku or ''
        ws.cell(row=row_num, column=3).value = book.title
        ws.cell(row=row_num, column=4).value = book.author
        ws.cell(row=row_num, column=5).value = book.author_oblozh
        ws.cell(row=row_num, column=6).value = book.genre
        ws.cell(row=row_num, column=7).value = book.publisher
        ws.cell(row=row_num, column=8).value = book.series or ''
        ws.cell(row=row_num, column=9).value = book.publication_year or ''
        ws.cell(row=row_num, column=10).value = book.language
        # Сохранность (преобразуем код в читаемое значение)
        ws.cell(row=row_num, column=11).value = book.get_condition_display() if book.condition else ''
        # Тип переплёта (преобразуем код в читаемое значение)
        ws.cell(row=row_num, column=12).value = book.get_cover_type_display()
        ws.cell(row=row_num, column=13).value = book.pages or ''
        ws.cell(row=row_num, column=14).value = book.isbn or ''
        # Цены (преобразуем Decimal в float)
        ws.cell(row=row_num, column=15).value = float(book.price) if book.price else 0
        ws.cell(row=row_num, column=16).value = float(book.old_price) if book.old_price else ''
        # НДС (преобразуем код в читаемое значение)
        ws.cell(row=row_num, column=17).value = book.get_vat_rate_display()
        ws.cell(row=row_num, column=18).value = book.stock
        # Размеры и вес (преобразуем Decimal в float)
        ws.cell(row=row_num, column=19).value = float(book.weight) if book.weight else ''
        ws.cell(row=row_num, column=20).value = float(book.length) if book.length else ''
        ws.cell(row=row_num, column=21).value = float(book.width) if book.width else ''
        ws.cell(row=row_num, column=22).value = float(book.height) if book.height else ''
        # Категория (название, если есть)
        ws.cell(row=row_num, column=23).value = book.category.name if book.category else ''
        # Источник (преобразуем код в читаемое значение)
        ws.cell(row=row_num, column=24).value = book.get_source_display()
        # Дата создания (форматируем)
        ws.cell(row=row_num, column=25).value = book.created_at.strftime('%Y-%m-%d %H:%M:%S')
    # Автоматическая ширина колонок на основе содержимого
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        # Находим максимальную длину текста в колонке
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Устанавливаем ширину с небольшим запасом, но не более 50 символов
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    # Создаем HTTP ответ с Excel файлом
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="books_export.xlsx"'
    # Сохраняем книгу Excel в ответ
    wb.save(response)
    return response
